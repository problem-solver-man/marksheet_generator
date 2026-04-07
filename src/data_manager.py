from config_manager import config_dir
from openpyxl import load_workbook
import shutil

# -------------------------
# LOAD RAW EXCEL DATA
# -------------------------

# Generator function to provide single worksheet at a time
def load_data(excel_path, selected_class):
    # Copy the original excel file into temporary file
    temp_excel_path = f"{config_dir}/temp_{excel_path.name}"
    shutil.copy(excel_path, temp_excel_path)

    wb = load_workbook(temp_excel_path, read_only=True, data_only=True)

    if selected_class == "All Class":
        for ws in wb.worksheets:
            yield ws
    else:
        yield wb[selected_class]



# -----------------------------
# RAW DATA  →  REQUIRED VALUES
# -----------------------------

# Function to prepare marksheet values for all students of a class
def prepare_data(ws, marksheet_order) -> list:
    empty_name_indicator = 0
    empty_sheet_indicator = True
    clas = ws.title.replace("Class-", "")
    wslist = []     # Worksheet-as-a-list

    for row in ws.iter_rows(min_row=2, max_col=50, values_only=True):
        row = list(row)
        
        # Stop if student name is empty in 10 continuous rows
        if row[1] is None:      # Name must be provided
            if empty_name_indicator >= 9:
                if empty_sheet_indicator:
                    return wslist   # Return the empty list
                break
            empty_name_indicator += 1
            continue            # Skip further processes

        empty_name_indicator = 0
        empty_sheet_indicator = False

        # Handle empty values
        for i in range(len(row)):
            if row[i] is None:
                row[i] = 0

        wslist.append([row[:2]+[clas], row[2:9], row[9:16], row[16:23],
                       row[23:30], row[30:37], row[37:44], row[44:]])

    # Add different total marks in each student list
    for student in wslist:
        # Total secured subjective marks (tssm)
        tssm = [sum(x for x in student[i][:2]
                    if isinstance(x, (int, float))) for i in range(1, 7)]
        tssm.append(student[7][0])

        # Failed subjects and failed by marks
        fails = [x for x in tssm if x < 30]
        failed_subjects = len(fails)
        # failed_by_marks = sum(30 - x for x in fails)

        # Pass up to -3 marks out of 210 marks (30 marks x 7 subjects)
        # if failed_by_marks <= 3:
        #     tssm = [max(x, 30) for x in tssm]
        #     failed_subjects = 0

        # Total secured marks (tsm)
        tsm = [sum(x for x in student[i]
                   if isinstance(x, (int, float))) for i in range(1, 8)]

        # Calculate total marks, percentage, and result (Pass or Fail)
        sum_tssm = sum(tssm)
        sum_tsm = sum(tsm)

        is_pass = failed_subjects == 0 and round(sum_tssm) >= 280
        result = "Pass" if is_pass else "Fail"
        percentage = f"{sum_tsm / 7:.2f}%"

        # Append [tssm], [tsm], and [sum of tssm, tsm] into wslist
        student.append(tssm)
        student.append(tsm)
        student.append([failed_subjects, sum_tssm, sum_tsm, result, percentage])

    # Sort students based on failed subjects, then by tsm, then by tssm
    wslist.sort(key=lambda x: (x[10][0], -x[10][2], -x[10][1]))

    # Add the rank of each student
    for index, student in enumerate(wslist):
        index += 1

        if 10 <= index % 100 <= 13:
            suffix = "th"
        else:
            suffix = {1: "st", 2: "nd", 3: "rd"}.get(index % 10, "th")

        student[10].append(f"{index}<sup>{suffix}</sup>")

    # Check the required marksheet order
    if marksheet_order == "Order by Rank":
        return wslist   # Return rank wise sorted student list

    # Sort students based on roll number
    wslist.sort(key=lambda x: x[0][0])
    return wslist       # Return roll wise sorted student list



# -------------------------
# CONVERT VALUES TO STRING
# -------------------------

# Function to convert all values to type str
def polish_data(student) -> list:
    tssm = student[8]

    for index, value in enumerate(tssm):
        if isinstance(value, (int, float)) and value < 30:
            tssm[index] = "*" + str(round(value))

    for item in student:
        for index, value in enumerate(item):
            if isinstance(value, float):
                item[index] = str(round(value))
            elif not isinstance(value, str):
                item[index] = str(value)

    return student
